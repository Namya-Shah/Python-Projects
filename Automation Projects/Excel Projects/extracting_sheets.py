import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('') # Enter filename with which you want to extract sheets

# Select the sheet by sheet number
sheet_number = 7  # Update with the desired sheet number
sheet = workbook.worksheets[sheet_number - 1]

# Extract numbers from the sheet
numbers = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]

# Create a new workbook
new_workbook = openpyxl.Workbook()

# Iterate over the numbers and create new sheets
while numbers:
    # Create a new sheet in the new workbook
    new_sheet = new_workbook.create_sheet()

    # Extract 20 numbers from the list
    extracted_numbers = numbers[:20]

    # Remove the extracted numbers from the list
    numbers = numbers[20:]

    # Write the extracted numbers to the new sheet
    for i, number in enumerate(extracted_numbers):
        new_sheet.cell(row=i+1, column=1, value=number)

# Save the new workbook
new_workbook.save('') # Enter your filename with which you want to save
